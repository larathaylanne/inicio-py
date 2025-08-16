from openpyxl import load_workbook, Workbook


planilha_country = load_workbook('Country AUTOMACAO.xlsx')
pagina_principal = planilha_country['Cereais']

planilha_respondida = Workbook()
aba_nova = planilha_respondida.active
aba_nova.tittle = 'Cereais'

for linha in pagina_principal.iter_rows(values_only=True):
    aba_nova.append(linha)

with open ('precos-country.txt', 'r', encoding='utf-8') as arquivo:
    for linha in arquivo:
        dados = linha.strip().split(',')
        aba_nova.append(dados)
planilha_respondida.save('preenchendo_planilha.xlsx')